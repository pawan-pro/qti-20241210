import pandas as pd
import datetime
import os
import json
import pickle
import re
import time
from google.auth.transport.requests import Request
from google.oauth2 import service_account
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# --- Configuration ---
SCOPES = ['https://www.googleapis.com/auth/calendar']
CALENDAR_ID = 'jatalep2018@email.iimcal.ac.in'
EXCEL_FILE = 'My_Calendar_Planner.xlsx'

# --- Authentication ---
def get_calendar_service():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    token_path = os.path.join(script_dir, 'token.pickle')
    creds_path = os.path.join(script_dir, 'credentials.json')

    creds = None
    
    if not os.path.exists(creds_path):
        print(f"Error: credentials.json not found at {creds_path}")
        return None

    with open(creds_path, 'r') as f:
        creds_data = json.load(f)
    
    if creds_data.get('type') == 'service_account':
        creds = service_account.Credentials.from_service_account_file(
            creds_path, scopes=SCOPES)
    else:
        if os.path.exists(token_path):
            with open(token_path, 'rb') as token:
                creds = pickle.load(token)
        
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    creds_path, SCOPES)
                creds = flow.run_local_server(port=0)
            
            with open(token_path, 'wb') as token:
                pickle.dump(creds, token)

    return build('calendar', 'v3', credentials=creds)

# --- Sync Up (Excel -> Google) ---
def sync_up_from_excel(service):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, EXCEL_FILE)

    if not os.path.exists(excel_path):
        print("No existing Excel file found. Skipping sync-up.")
        return

    print(f"Reading changes from {EXCEL_FILE}...")
    try:
        df_planner = pd.read_excel(excel_path, sheet_name='Planner', index_col=0)
        df_meta = pd.read_excel(excel_path, sheet_name='Metadata', index_col=0)
    except Exception as e:
        print(f"Error reading Excel: {e}")
        print("Make sure the file is not open in another program.")
        return

    changes_count = 0

    for date_idx, row in df_planner.iterrows():
        for col_name in df_planner.columns:
            if col_name == 'Day': continue
            
            cell_value = str(row[col_name]) if pd.notna(row[col_name]) else ""
            cell_value = cell_value.strip()
            
            try:
                meta_value = str(df_meta.loc[date_idx, col_name]) if pd.notna(df_meta.loc[date_idx, col_name]) else ""
            except KeyError:
                meta_value = ""
            meta_value = meta_value.strip()

            if not cell_value and not meta_value:
                continue

            # Parse Metadata
            meta_entries = []
            if meta_value:
                parts = meta_value.split('||')
                for p in parts:
                    if '|' in p:
                        eid, esum = p.split('|', 1)
                        meta_entries.append({'id': eid, 'summary': esum})

            # Parse Cell Lines
            cell_lines = [line.strip() for line in cell_value.split('\n') if line.strip()]

            # --- Logic ---
            
            # 1. DELETE: Meta exists, Cell is empty
            if meta_entries and not cell_lines:
                for entry in meta_entries:
                    print(f"Deleting event: {entry['summary']}")
                    try:
                        service.events().delete(calendarId=CALENDAR_ID, eventId=entry['id']).execute()
                        changes_count += 1
                    except HttpError as e:
                        if e.resp.status == 410: pass # Already deleted
                        else: print(f"Error deleting: {e}")

            # 2. CREATE: No Meta, Cell has text
            elif not meta_entries and cell_lines:
                for line in cell_lines:
                    # Parse time from column + date
                    date_str = pd.to_datetime(date_idx).strftime('%Y-%m-%d')
                    hour_str = col_name # "10:00"
                    
                    # Heuristic: If line starts with "10:00 ", strip it for summary
                    time_pattern = re.compile(r'^\d{1,2}:\d{2}\s+')
                    summary = time_pattern.sub('', line).strip()
                    
                    start_dt_str = f"{date_str}T{hour_str}:00"
                    start_dt = datetime.datetime.strptime(start_dt_str, '%Y-%m-%dT%H:%M:%S')
                    end_dt = start_dt + datetime.timedelta(hours=1)
                    
                    print(f"Creating event: {date_str} {hour_str} - {summary}")
                    event_body = {
                        'summary': summary,
                        'start': {'dateTime': start_dt.isoformat(), 'timeZone': 'Asia/Kolkata'},
                        'end': {'dateTime': end_dt.isoformat(), 'timeZone': 'Asia/Kolkata'}
                    }
                    try:
                        service.events().insert(calendarId=CALENDAR_ID, body=event_body).execute()
                        changes_count += 1
                    except HttpError as e:
                        print(f"Error creating: {e}")

            # 3. UPDATE: 1 Meta, 1 Cell Line (Simple Case)
            elif len(meta_entries) == 1 and len(cell_lines) == 1:
                meta_entry = meta_entries[0]
                current_text = cell_lines[0]
                
                time_pattern = re.compile(r'^\d{1,2}:\d{2}\s+')
                clean_summary = time_pattern.sub('', current_text).strip()
                
                if clean_summary != meta_entry['summary']:
                    print(f"Updating event: {meta_entry['summary']} -> {clean_summary}")
                    try:
                        event = service.events().get(calendarId=CALENDAR_ID, eventId=meta_entry['id']).execute()
                        event['summary'] = clean_summary
                        service.events().update(calendarId=CALENDAR_ID, eventId=meta_entry['id'], body=event).execute()
                        changes_count += 1
                    except HttpError as e:
                        print(f"Error updating: {e}")

    if changes_count > 0:
        print(f"Synced {changes_count} changes to Google Calendar.")
        # Wait a moment for propagation
        time.sleep(2)
    else:
        print("No changes detected in Excel.")

# --- Sync Down (Google -> Excel) ---
def sync_down_to_excel(service):
    print("Downloading latest calendar data...")
    
    # Range: Past 30 days to Next 90 days
    now = datetime.datetime.now(datetime.timezone.utc)
    time_min = (now - datetime.timedelta(days=30)).isoformat()
    
    events_result = service.events().list(
        calendarId=CALENDAR_ID, timeMin=time_min,
        maxResults=500, singleEvents=True,
        orderBy='startTime').execute()
    events = events_result.get('items', [])

    # Determine Date Range for Grid
    today = datetime.datetime.now()
    first_day_this_month = today.replace(day=1)
    last_month = first_day_this_month - datetime.timedelta(days=1)
    start_date = last_month.replace(day=1)
    
    next_month = (first_day_this_month + datetime.timedelta(days=32)).replace(day=1)
    next_next_month = (next_month + datetime.timedelta(days=32)).replace(day=1)
    end_date = next_next_month - datetime.timedelta(days=1)
    
    # Create Grid
    date_range = pd.date_range(start=start_date, end=end_date)
    hours = [f"{h:02d}:00" for h in range(24)]
    
    df_display = pd.DataFrame('', index=date_range, columns=hours)
    df_meta = pd.DataFrame('', index=date_range, columns=hours)

    # Populate Grid
    for event in events:
        start = event['start'].get('dateTime', event['start'].get('date'))
        summary = event.get('summary', '(No Title)')
        eid = event['id']
        
        # Parse Start Time
        try:
            if 'T' in start:
                dt = datetime.datetime.fromisoformat(start)
            else:
                dt = datetime.datetime.strptime(start, '%Y-%m-%d')
                # For all day events, we might skip or put in 00:00
                # Let's put in 00:00 for now
            
            # Convert to naive for comparison with pandas index (which we'll make naive)
            # Actually, let's just use string matching for the date
            row_idx = dt.strftime('%Y-%m-%d')
            hour_col = f"{dt.hour:02d}:00"
            
            # Check if in range
            if start_date.strftime('%Y-%m-%d') <= row_idx <= end_date.strftime('%Y-%m-%d'):
                # Display Text
                display_text = f"{dt.strftime('%H:%M')} {summary}"
                
                # Metadata Text
                meta_text = f"{eid}|{summary}"
                
                # Add to Display DF
                try:
                    # We access by date object (naive)
                    d_date = dt.date()
                    current_disp = df_display.loc[d_date, hour_col]
                    if current_disp:
                        df_display.loc[d_date, hour_col] = current_disp + "\n" + display_text
                    else:
                        df_display.loc[d_date, hour_col] = display_text
                        
                    # Add to Meta DF
                    current_meta = df_meta.loc[d_date, hour_col]
                    if current_meta:
                        df_meta.loc[d_date, hour_col] = current_meta + "||" + meta_text
                    else:
                        df_meta.loc[d_date, hour_col] = meta_text
                except KeyError:
                    pass
        except ValueError:
            continue

    # Formatting
    df_display.insert(0, 'Day', df_display.index.day_name())
    df_display.index.name = 'Date'
    df_display.index = df_display.index.tz_localize(None) # Make naive for Excel
    
    df_meta.index = df_meta.index.tz_localize(None)

    # Save
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, EXCEL_FILE)
    
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_display.to_excel(writer, sheet_name='Planner')
            df_meta.to_excel(writer, sheet_name='Metadata')
    except PermissionError:
        print(f"CRITICAL ERROR: Could not save {EXCEL_FILE}. Is it open? Please close it and run again.")
        return

    # Apply Styles
    wb = load_workbook(excel_path)
    ws = wb['Planner']
    
    # Hide Metadata
    if 'Metadata' in wb.sheetnames:
        wb['Metadata'].sheet_state = 'hidden'

    # Styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    weekend_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    event_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        day_cell = row[1]
        is_weekend = day_cell.value in ['Saturday', 'Sunday']
        
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if is_weekend: cell.fill = weekend_fill
            if cell.col_idx > 2 and cell.value: cell.fill = event_fill

    # Dimensions
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    for col in range(3, 27):
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[col_letter].width = 20
    
    ws.freeze_panes = 'C2'
    wb.save(excel_path)
    print(f"Successfully updated {EXCEL_FILE}")

def main():
    service = get_calendar_service()
    if not service: return
    
    # 1. Sync Up (Push changes from Excel to Google)
    sync_up_from_excel(service)
    
    # 2. Sync Down (Pull latest from Google to Excel)
    sync_down_to_excel(service)

if __name__ == "__main__":
    main()
